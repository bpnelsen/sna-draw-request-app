"""
Vercel Python Serverless Function
Processes SNA Draw Request Excel files
"""

import os
import sys
import base64
import json
import tempfile
from pathlib import Path

# Add openpyxl to requirements if not present
try:
    import openpyxl
except ImportError:
    # In serverless, dependencies come from requirements.txt
    raise ImportError("openpyxl not installed. Add to requirements.txt: openpyxl pandas")

from openpyxl.styles import Font, PatternFill
from collections import defaultdict

def reorganize_excel(input_file, output_file):
    """Process and reorganize Excel file by SN Loan #"""
    try:
        # Load workbook
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        # Read headers
        headers = {}
        data_by_lot = defaultdict(list)
        
        for col_num, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[col_num] = cell.value
        
        # Group by SN Loan # (Column G)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[6].value:  # Column G
                lot_name = str(row[6].value).strip()
                data_by_lot[lot_name].append(row)
        
        # Create output workbook
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = "Reorganized"
        
        # Write headers
        for col_num, header in headers.items():
            ws_output.cell(row=1, column=col_num, value=header)
        
        # Write data grouped by lot
        current_row = 2
        for lot_name in sorted(data_by_lot.keys()):
            for row_data in data_by_lot[lot_name]:
                for col_num, cell in enumerate(row_data, 1):
                    ws_output.cell(row=current_row, column=col_num, value=cell.value)
                current_row += 1
        
        # Save output
        wb_output.save(output_file)
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False

async def handler(request):
    """Handle file upload and processing"""
    
    if request.method != 'POST':
        return {'statusCode': 405, 'body': json.dumps({'error': 'Method not allowed'})}
    
    try:
        # Parse form data
        content_type = request.headers.get('content-type', '')
        
        if 'multipart/form-data' not in content_type:
            return {
                'statusCode': 400,
                'body': json.dumps({'error': 'Expected multipart/form-data'})
            }
        
        # Get file from request
        if 'file' not in request.files:
            return {
                'statusCode': 400,
                'body': json.dumps({'error': 'No file provided'})
            }
        
        file = request.files['file']
        filename = file.filename
        
        # Validate
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return {
                'statusCode': 400,
                'body': json.dumps({'error': 'Only Excel files supported'})
            }
        
        # Process in temp directory
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / filename
            output_name = f"organized_{filename}"
            output_path = Path(tmpdir) / output_name
            
            # Save uploaded file
            file.save(str(input_path))
            
            # Process
            if not reorganize_excel(str(input_path), str(output_path)):
                return {
                    'statusCode': 500,
                    'body': json.dumps({'error': 'Processing failed'})
                }
            
            # Read and encode result
            with open(output_path, 'rb') as f:
                processed_data = base64.b64encode(f.read()).decode()
            
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'success': True,
                    'fileName': output_name,
                    'processedData': processed_data,
                    'message': 'File processed successfully!'
                })
            }
            
    except Exception as e:
        print(f"Error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': 'Internal server error', 'details': str(e)})
        }
